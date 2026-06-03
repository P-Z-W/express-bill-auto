# -*- coding: utf-8 -*-
"""
快递单合并模块（V3.0）
==================================================
【框架说明】
  功能：读取 data/YYYY-MM/ 目录下各快递公司原始账单
       → 清洗格式 → 统一字段 → 合并成一张总表
       → 输出到 output/YYYY-MM/清洗合并总账单.xlsx
  结构：工具函数 → 数据清洗 → 快递识别 → 合并 → 导出美化

【V3.0 核心升级】
  - 从 config/express_config.json 读取快递公司配置
  - 根据配置的 identify_column 字段自动识别快递类型
  - 支持任意快递公司（申通/中通/圆通/极兔等），无需改代码
  - enabled=false 的快递公司自动跳过
  - 保持对 V2.x 数据的完全兼容（字段结构不变）

【V2.4 架构（保留）】
  - 删除重复的 ensure_output_dir / read_excel_file / clean_dataframe
  - 改从 utils 导入统一工具函数和样式常量
  - 业务逻辑完全不变

【快递识别逻辑】
  扫描 data/YYYY-MM/ 目录所有 xlsx 文件，
  读取表头，按 express_config.json 中各快递的 identify_column 匹配
  例：含"业务时间"列 → 申通，含"扫描时间"列 → 中通

【可独立运行】
==================================================
"""
import pandas as pd
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from config import settings
from utils import ensure_folder, read_excel, clean_dataframe, FULL_BORDER

# ====================== 路径配置 ======================
DATA_FOLDER         = settings.DATA_FOLDER
OUTPUT_FOLDER       = settings.OUTPUT_FOLDER
OUTPUT_FILE         = os.path.join(OUTPUT_FOLDER, settings.EXPRESS_OUTPUT_FILE)
EXPRESS_CONFIG_PATH = os.path.join("config", "express_config.json")

# ====================== 统一输出字段 ======================
FINAL_COLUMNS = ["业务时间", "运单号", "目的省份", "目的城市", "结算重量", "快递类型", "所属团队"]


# ====================== 读取快递配置 ======================
def load_express_config():
    """
    从 config/express_config.json 读取已启用的快递公司配置
    返回：[{"name": "申通", "identify_column": "业务时间", "enabled": True}, ...]
    文件不存在时返回默认配置（申通+中通），保持向后兼容
    """
    default_config = [
        {"name": "申通", "identify_column": "业务时间", "enabled": True},
        {"name": "中通", "identify_column": "扫描时间", "enabled": True}
    ]
    try:
        with open(EXPRESS_CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        express_list = data.get("express_list", default_config)
        # 只返回已启用的快递公司
        enabled = [e for e in express_list if e.get("enabled", True)]
        if not enabled:
            print("⚠️  express_config.json 中所有快递均已禁用，使用默认配置")
            return default_config
        print(f"✅ 已读取快递配置，启用快递：{[e['name'] for e in enabled]}")
        return enabled
    except FileNotFoundError:
        print(f"⚠️  未找到 {EXPRESS_CONFIG_PATH}，使用默认配置（申通+中通）")
        return default_config
    except Exception as e:
        print(f"⚠️  读取快递配置失败：{str(e)}，使用默认配置")
        return default_config


# ====================== 自动扫描识别文件 ======================
def auto_find_express_files(data_folder, express_config):
    """
    自动扫描 data/YYYY-MM/ 目录，按 express_config 中的 identify_column 识别文件
    支持任意快递公司，只要在配置中定义了识别列

    参数：
      data_folder    - 账单目录路径
      express_config - 已启用的快递配置列表

    返回：
      dict，格式 {"申通": "/path/to/file.xlsx", "中通": "/path/to/file.xlsx"}
      未识别到的快递不在字典中
    """
    if not os.path.exists(data_folder):
        print(f"❌ 数据目录不存在：{data_folder}")
        return {}

    xlsx_files = [
        os.path.join(data_folder, f)
        for f in os.listdir(data_folder)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    if not xlsx_files:
        print(f"❌ {data_folder} 目录下没有找到任何xlsx文件")
        return {}

    # 构建识别列 → 快递名称的映射
    # 用于快速查找：{"业务时间": "申通", "扫描时间": "中通", ...}
    col_to_express = {
        e["identify_column"]: e["name"]
        for e in express_config
    }

    found = {}       # 已匹配的快递文件
    unmatched = []   # 未能识别的文件

    for fpath in xlsx_files:
        # 只读表头，速度快
        df_head = read_excel(fpath, nrows=0)
        if df_head is None:
            continue
        cols = df_head.columns.tolist()

        matched_express = None
        for col, express_name in col_to_express.items():
            if col in cols and express_name not in found:
                matched_express = express_name
                found[express_name] = fpath
                print(f"✅ 识别到{express_name}账单：{os.path.basename(fpath)}（识别列：{col}）")
                break

        if matched_express is None:
            unmatched.append(os.path.basename(fpath))

    if unmatched:
        print(f"⚠️  以下文件未匹配到任何快递配置，已跳过：{unmatched}")

    # 检查哪些已启用的快递没有找到文件
    for e in express_config:
        if e["name"] not in found:
            print(f"⚠️  未识别到{e['name']}账单（需含'{e['identify_column']}'列），请检查 {data_folder} 目录")

    return found


# ====================== 通用账单字段标准化 ======================
def standardize_dataframe(df, express_name, identify_column):
    """
    将各快递原始账单字段标准化为统一格式
    不同快递列名不同，这里做统一映射

    统一输出字段：业务时间 / 运单号 / 目的省份 / 目的城市 / 结算重量 / 快递类型 / 所属团队

    当前支持的快递映射规则：
      申通：业务时间, 运单号, 订单目的省份→目的省份, 订单目的城市→目的城市, 结算重量
      中通：扫描时间→业务时间, 运单号, 目的地→目的省份, 目的地→目的城市, 结算重量
      圆通/极兔：使用发货时间/揽收时间作为业务时间，其余字段需按实际情况补充
    """
    df = clean_dataframe(df)
    if df is None:
        return None

    cols = df.columns.tolist()
    result = {}

    # ---- 业务时间（时间字段，不同快递列名不同）----
    time_col_candidates = [identify_column, "业务时间", "扫描时间", "发货时间", "揽收时间", "下单时间"]
    time_col = next((c for c in time_col_candidates if c in cols), None)
    if time_col:
        result["业务时间"] = df[time_col]
    else:
        result["业务时间"] = ""
        print(f"  ⚠️  {express_name}：未找到时间列，业务时间置空")

    # ---- 运单号 ----
    waybill_candidates = ["运单号", "快递单号", "面单号", "waybill_id"]
    waybill_col = next((c for c in waybill_candidates if c in cols), None)
    if waybill_col:
        result["运单号"] = df[waybill_col].astype(str).str.strip()
    else:
        print(f"  ❌ {express_name}：未找到运单号列，该快递跳过")
        return None

    # ---- 目的省份 ----
    prov_candidates = ["订单目的省份", "目的省份", "目的地省份", "收件省", "目的地"]
    prov_col = next((c for c in prov_candidates if c in cols), None)
    result["目的省份"] = df[prov_col] if prov_col else ""

    # ---- 目的城市 ----
    city_candidates = ["订单目的城市", "目的城市", "目的地城市", "收件市", "目的地"]
    city_col = next((c for c in city_candidates if c in cols), None)
    result["目的城市"] = df[city_col] if city_col else ""

    # ---- 结算重量 ----
    weight_candidates = ["结算重量", "计费重量", "重量", "实重"]
    weight_col = next((c for c in weight_candidates if c in cols), None)
    if weight_col:
        result["结算重量"] = pd.to_numeric(df[weight_col], errors="coerce").fillna(0)
    else:
        result["结算重量"] = 0
        print(f"  ⚠️  {express_name}：未找到重量列，结算重量置0")

    # ---- 快递类型 & 所属团队 ----
    result["快递类型"] = express_name
    result["所属团队"] = ""

    standardized = pd.DataFrame(result)[FINAL_COLUMNS]
    standardized  = standardized.dropna(subset=["运单号"])
    print(f"  ✅ {express_name}：标准化完成，共 {len(standardized)} 条记录")
    return standardized


# ====================== 合并所有账单 ======================
def merge_all_bills(dfs_map):
    """
    合并所有已标准化的账单 DataFrame
    dfs_map: {"申通": df_st, "中通": df_zt, ...}
    支持只有部分快递有数据的情况
    """
    valid_dfs = [df for df in dfs_map.values() if df is not None and not df.empty]
    if not valid_dfs:
        raise ValueError("所有快递账单均为空或识别失败，无法合并")

    merged = pd.concat(valid_dfs, ignore_index=True)
    merged = merged[FINAL_COLUMNS]
    merged = clean_dataframe(merged)
    merged = merged.dropna(subset=["运单号"])

    # 统计汇总
    total = len(merged)
    for name, df in dfs_map.items():
        if df is not None and not df.empty:
            count = len(df)
            pct   = round(count / total * 100, 1) if total > 0 else 0
            print(f"   - {name}：{count} 条（{pct}%）")
    print(f"   合计：{total} 条")
    return merged


# ====================== 导出并美化 ======================
def save_and_style(df):
    """导出Excel并添加表头、边框、居中样式（使用utils统一样式）"""
    ensure_folder(OUTPUT_FOLDER)
    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # 表头加粗居中
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = FULL_BORDER

    # 数据行加边框
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = FULL_BORDER

    wb.save(OUTPUT_FILE)
    print(f"✅ 合并完成，文件已生成：{OUTPUT_FILE}")


# ====================== 主流程 ======================
def run_merge_process():
    """
    快递合并主流程（V3.0）
    1. 读取 express_config.json 获取已启用快递配置
    2. 扫描 data/YYYY-MM/ 按识别列自动匹配文件
    3. 统一标准化各快递字段
    4. 合并并导出
    """
    print("🚀 开始合并快递账单（V3.0 动态识别模式）...")

    # Step 1：读取快递配置
    express_config = load_express_config()

    # Step 2：扫描目录，自动识别文件
    found_files = auto_find_express_files(DATA_FOLDER, express_config)
    if not found_files:
        print(f"❌ 未识别到任何快递账单，请检查 {DATA_FOLDER} 目录")
        return None

    # Step 3：读取并标准化每个快递账单
    dfs_map = {}
    for express in express_config:
        name            = express["name"]
        identify_column = express["identify_column"]
        if name not in found_files:
            continue
        fpath = found_files[name]
        print(f"\n📋 处理{name}账单：{os.path.basename(fpath)}")
        df_raw = read_excel(fpath)
        if df_raw is None:
            print(f"  ❌ 读取失败，跳过{name}")
            continue
        df_std = standardize_dataframe(df_raw, name, identify_column)
        dfs_map[name] = df_std

    if not dfs_map:
        print("❌ 所有快递账单处理失败")
        return None

    # Step 4：合并
    print(f"\n📊 合并账单统计：")
    df_all = merge_all_bills(dfs_map)

    # Step 5：导出
    save_and_style(df_all)
    return df_all


if __name__ == "__main__":
    run_merge_process()