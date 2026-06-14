# -*- coding: utf-8 -*-
"""
快递订单对账匹配模块（V3.3）
==================================================
【功能说明】
  1. 运单号匹配 → 回填所属团队
  2. 智能判断计费模式 → 按结算重量+目的省份自动生成「实际计算方式」
  3. 精准计算应付金额 → 读取config/price_config.xlsx
  4. 格式化导出结果 → 带边框、表头样式、固定列宽的Excel
  5. 完善容错机制 → 缺失文件/配置时给出明确报错

【计费规则】
  1. 全国均重：结算重量≤1kg 或 1-3kg非新疆西藏 → 金额留空
  2. 单票：结算重量≥3kg → 重量×续重单价 + (超3kg面单费 - 充单价格)
  3. 新西1-3公斤：新疆/西藏 且 1kg＜重量＜3kg → 3kg内面单费 + 重量×续重单价 - 充单价格

【V2.4 架构整理】
  - 删除重复的 ensure_folder 函数
  - 改从 utils 导入统一工具函数和样式常量
  - 业务逻辑、计费规则完全不变

【price_config.xlsx字段约定】
  申通/中通报价：A列=省份 B列=3kg内面单费 D列=超3kg面单费 E列=续重单价
  客户快递加收单价信息记录：K列=快递类型 L列=充单价格
==================================================
"""
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import settings
from core.utils import ensure_folder, read_excel, FULL_BORDER, HEADER_FONT, CENTER_ALIGN, BOLD_FONT

# ====================== 路径配置 ======================
OUTPUT_FOLDER     = settings.OUTPUT_FOLDER
EXPRESS_FILE      = os.path.join(OUTPUT_FOLDER, settings.EXPRESS_OUTPUT_FILE)
RESULT_FILE       = os.path.join(OUTPUT_FOLDER, settings.RESULT_FILE)
PRICE_CONFIG_PATH = os.path.join(settings.CONFIG_FOLDER, "price_config.xlsx")


# ====================== 计费模式判断 ======================
def get_calc_type(dest_prov, weight):
    """
    根据目的省份+结算重量自动判断计费模式
    单票：重量≥3
    新西1-3公斤：新疆/西藏 且 1kg＜重量＜3kg
    其余：全国均重
    """
    if weight >= 3:
        return "单票"
    if (dest_prov.startswith("新疆") or dest_prov.startswith("西藏")) and 1 < weight < 3:
        return "新西1-3公斤"
    return "全国均重"


# ====================== 报价配置加载 ======================
def load_price_config():
    """读取申通、中通报价 + 充单价格"""
    price_dict = {}

    # 读取申通报价
    df_st = read_excel(PRICE_CONFIG_PATH)
    if df_st is None:
        raise RuntimeError("读取申通报价失败")

    # 重新按sheet读取
    try:
        df_st = pd.read_excel(PRICE_CONFIG_PATH, sheet_name="申通报价")
    except Exception as e:
        raise RuntimeError(f"读取申通报价失败：{str(e)}") from e

    st_map = {}
    for idx, row in df_st.iterrows():
        if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
            continue
        prov_name = str(row.iloc[0]).strip()
        try:
            st_map[prov_name] = {
                "面单费_3kg内": float(row.iloc[1]),
                "面单费_超3kg": float(row.iloc[3]),
                "续重单价":     float(row.iloc[4])
            }
        except Exception as e:
            raise RuntimeError(f"申通报价第{idx+2}行数据异常：{str(e)}") from e
    price_dict["申通"] = st_map

    # 读取中通报价
    try:
        df_zt = pd.read_excel(PRICE_CONFIG_PATH, sheet_name="中通报价")
    except Exception as e:
        raise RuntimeError(f"读取中通报价失败：{str(e)}") from e

    zt_map = {}
    for idx, row in df_zt.iterrows():
        if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
            continue
        prov_name = str(row.iloc[0]).strip()
        try:
            zt_map[prov_name] = {
                "面单费_3kg内": float(row.iloc[1]),
                "面单费_超3kg": float(row.iloc[3]),
                "续重单价":     float(row.iloc[4])
            }
        except Exception as e:
            raise RuntimeError(f"中通报价第{idx+2}行数据异常：{str(e)}") from e
    price_dict["中通"] = zt_map

    # 读取充单价格
    try:
        df_charge = pd.read_excel(PRICE_CONFIG_PATH, sheet_name="客户快递加收单价信息记录")
    except Exception as e:
        raise RuntimeError(f"读取充单价格失败：{str(e)}") from e

    charge_map = {}
    for idx, row in df_charge.iterrows():
        if pd.isna(row.iloc[10]) or pd.isna(row.iloc[11]):
            continue
        type_name = str(row.iloc[10]).strip()
        if type_name in ["申通", "中通"]:
            try:
                charge_map[type_name] = float(row.iloc[11])
            except Exception as e:
                raise RuntimeError(f"充单价格第{idx+2}行数据异常：{str(e)}") from e

    if "申通" not in charge_map or "中通" not in charge_map:
        raise RuntimeError("充单价格配置不全，请检查 price_config.xlsx")

    price_dict["充单价"] = charge_map
    print(f"✅ 报价表加载完成：申通{len(st_map)}省份，中通{len(zt_map)}省份，充单价：{charge_map}")
    return price_dict


# ====================== 应付金额计算 ======================
def calculate_single_fee(row, price_dict):
    """
    按计费模式计算单条订单应付金额
    智能省份匹配：开头模糊匹配，兼容带省/自治区后缀
    """
    calc_method  = row["实际计算方式"]
    express_type = str(row["快递类型"]).strip()
    dest_prov    = str(row["目的省份"]).strip()

    try:
        weight = float(row["结算重量"])
    except Exception:
        return round(0.00, 2)

    if calc_method == "全国均重":
        return ""

    if express_type not in ["申通", "中通"]:
        return round(0.00, 2)

    charge_price     = price_dict["充单价"][express_type]
    express_prov_map = price_dict[express_type]

    # 智能开头匹配省份
    match_prov = None
    for p_key in express_prov_map:
        if dest_prov.startswith(p_key):
            match_prov = p_key
            break

    if not match_prov:
        print(f"⚠️  运单号{row['运单号']}：{express_type}无'{dest_prov}'省份报价，金额置0")
        return round(0.00, 2)

    prov_data = express_prov_map[match_prov]

    if calc_method == "单票":
        fee = weight * prov_data["续重单价"] + (prov_data["面单费_超3kg"] - charge_price)
    elif calc_method == "新西1-3公斤":
        fee = prov_data["面单费_3kg内"] + weight * prov_data["续重单价"] - charge_price
    else:
        return round(0.00, 2)

    return round(fee, 2)


def calculate_fees_vectorized(df, price_dict):
    """向量化计算应付金额，逐元素 round(2) 与行级计算完全一致"""
    result = pd.Series("", index=df.index, dtype=object)

    for express_type in ["申通", "中通"]:
        charge_price = price_dict["充单价"][express_type]
        prov_map     = price_dict[express_type]

        for calc_method in ["单票", "新西1-3公斤"]:
            mask = (
                (df["快递类型"].astype(str).str.strip() == express_type) &
                (df["实际计算方式"] == calc_method)
            )
            if not mask.any():
                continue

            sub        = df[mask]
            dest_provs = sub["目的省份"].astype(str).str.strip()

            # 与原行级逻辑一致：取第一个 startswith 命中的 key
            unique_provs = dest_provs.unique()
            prov_to_key  = {}
            for p in unique_provs:
                for k in prov_map:
                    if p.startswith(k):
                        prov_to_key[p] = k
                        break

            matched_keys = dest_provs.map(prov_to_key)

            no_match = matched_keys.isna()
            if no_match.any():
                for idx in sub[no_match].index:
                    print(f"⚠️  运单号{df.loc[idx, '运单号']}：{express_type}无'{df.loc[idx, '目的省份']}'省份报价，金额置0")
                result[sub[no_match].index] = round(0.00, 2)

            ok = ~no_match
            if not ok.any():
                continue

            sub_ok  = sub[ok]
            keys_ok = matched_keys[ok]
            weights = pd.to_numeric(sub_ok["结算重量"], errors="coerce")

            bad_w = weights.isna()
            if bad_w.any():
                result[sub_ok[bad_w].index] = round(0.00, 2)

            good = ~bad_w
            if not good.any():
                continue

            sub_g   = sub_ok[good]
            keys_g  = keys_ok[good]
            w_g     = weights[good]

            xuchong     = keys_g.map(lambda k: prov_map[k]["续重单价"])
            fee_over3kg = keys_g.map(lambda k: prov_map[k]["面单费_超3kg"])
            fee_3kg_in  = keys_g.map(lambda k: prov_map[k]["面单费_3kg内"])

            if calc_method == "单票":
                fees = (w_g * xuchong + (fee_over3kg - charge_price)).round(2)
            else:
                fees = (fee_3kg_in + w_g * xuchong - charge_price).round(2)

            result[sub_g.index] = fees

    mask_other = (
        ~df["快递类型"].astype(str).str.strip().isin(["申通", "中通"]) &
        (df["实际计算方式"] != "全国均重")
    )
    result[mask_other] = round(0.00, 2)

    return result


# ====================== 源数据读取 ======================
def load_source_data():
    """读取清洗合并账单 + 订单匹配文件"""
    # 读取清洗合并总账单
    if not os.path.exists(EXPRESS_FILE):
        raise RuntimeError("未找到清洗合并总账单，请先运行 merge_express.py")

    df_express = read_excel(EXPRESS_FILE)
    if df_express is None:
        raise RuntimeError("清洗合并账单读取失败")

    required_cols = ["运单号", "目的省份", "结算重量", "快递类型"]
    missing = [c for c in required_cols if c not in df_express.columns]
    if missing:
        raise RuntimeError(f"清洗合并账单缺少关键列：{missing}")
    print(f"✅ 成功读取清洗合并总账单：共 {len(df_express)} 条记录")

    # 查找订单文件
    order_files = [
        f for f in os.listdir(OUTPUT_FOLDER)
        if f.startswith(settings.ORDER_FILE_PREFIX)
    ]
    if not order_files:
        raise RuntimeError("未找到订单匹配文件，请先运行 order_db.py")

    latest_order_file = sorted(order_files)[-1]
    order_file_path   = os.path.join(OUTPUT_FOLDER, latest_order_file)
    df_order = read_excel(order_file_path)
    if df_order is None:
        raise RuntimeError("订单匹配文件读取失败")

    required_cols_order = ["运单号", "所属团队"]
    missing_order = [c for c in required_cols_order if c not in df_order.columns]
    if missing_order:
        raise RuntimeError(f"订单匹配文件缺少关键列：{missing_order}")
    print(f"✅ 成功读取订单匹配文件：{latest_order_file}，共 {len(df_order)} 条记录")

    return df_express, df_order


# ====================== 运单号匹配 ======================
def match_team_by_waybill():
    """核心业务：加载 → 清洗运单号 → 匹配团队 → 生成计算方式 → 算金额"""
    df_express, df_order = load_source_data()

    # 运单号清洗
    df_express["运单号"] = df_express["运单号"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df_order["运单号"]   = df_order["运单号"].astype(str).str.strip()
    print("✅ 运单号清洗完成")

    # 构建映射字典
    waybill_team_map = df_order[["运单号", "所属团队"]].drop_duplicates(subset="运单号")\
        .set_index("运单号")["所属团队"].to_dict()
    print(f"✅ 构建运单号-团队映射：共 {len(waybill_team_map)} 个唯一运单号")

    # 回填所属团队
    df_express["所属团队"] = df_express["运单号"].map(waybill_team_map).fillna("未匹配")
    df_express["实际计算方式"] = ""
    df_express["单票应付金额"] = ""

    # 批量判断计费模式
    print("🔄 正在批量判断计费模式...")
    _prov   = df_express["目的省份"].astype(str).str.strip()
    _weight = pd.to_numeric(df_express["结算重量"], errors="coerce")
    _xinxi  = _prov.str.startswith("新疆") | _prov.str.startswith("西藏")
    _xinxi_13 = _xinxi & (_weight > 1) & (_weight < 3)
    df_express["实际计算方式"] = np.select(
        [_weight >= 3, _xinxi_13],
        ["单票", "新西1-3公斤"],
        default="全国均重"
    )
    method_dist = df_express["实际计算方式"].value_counts()
    for method, count in method_dist.items():
        print(f"   - {method}：{count} 条（{count/len(df_express)*100:.1f}%）")

    # 批量计算金额
    print("🔄 正在加载报价表并计算应付金额...")
    price_dict = load_price_config()
    df_express["单票应付金额"] = calculate_fees_vectorized(df_express, price_dict)

    matched   = len(df_express[df_express["所属团队"] != "未匹配"])
    unmatched = len(df_express) - matched
    print(f"✅ 匹配完成：总{len(df_express)}条，已匹配{matched}条，未匹配{unmatched}条")

    return df_express


# ====================== 结果导出 ======================
def export_styled_result(df_result):
    """导出带样式Excel（使用utils统一样式常量）"""
    ensure_folder(OUTPUT_FOLDER)
    df_result.to_excel(RESULT_FILE, index=False, engine="openpyxl")

    wb = load_workbook(RESULT_FILE)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 全表边框+居中
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border    = FULL_BORDER
            cell.alignment = CENTER_ALIGN

    # 表头样式（蓝色背景+白色加粗）
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # 固定列宽
    column_widths = {
        'A': 20, 'B': 18, 'C': 12, 'D': 12,
        'E': 10, 'F': 16, 'G': 16, 'H': 16, 'I': 28
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(RESULT_FILE)
    print(f"✅ 最终对账结果已导出：{RESULT_FILE}")


# ====================== 主函数 ======================
def run_reconciliation():
    """程序主入口"""
    print("=" * 70)
    print("📦 快递订单对账匹配程序（V2.4）")
    print(f"📂 输出路径：{OUTPUT_FOLDER}")
    print("=" * 70)

    try:
        matched_data = match_team_by_waybill()
    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")
        raise

    if matched_data is not None and len(matched_data) > 0:
        export_styled_result(matched_data)
        print(f"\n🎉 对账完成 → {RESULT_FILE}")
    else:
        print("\n⚠️  无有效对账数据")

    print("\n" + "=" * 70)


if __name__ == "__main__":
    run_reconciliation()
