# -*- coding: utf-8 -*-
"""
数据库订单下载模块（V2.4 稳定版）
==================================================
【框架说明】
  功能：从阿里云MySQL下载订单数据
       → 读取外部SQL语句（动态日期范围）
       → 导出Excel并高速美化
       → 保存到 output/YYYY-MM/毅播快递数据_YYYYMM.xlsx

【V2.4 架构整理】
  - 删除重复的 ensure_folder / get_last_month_str 函数
  - 改从 utils 导入统一工具函数和样式常量
  - 业务逻辑完全不变

【可独立运行】
==================================================
"""
import pymysql
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from config import settings
from core.utils import ensure_folder, get_last_month_str, FULL_BORDER, HEADER_FONT, CENTER_ALIGN

# ====================== 配置（从settings读取）======================
DB_CONFIG         = settings.DB_CONFIG
OUTPUT_FOLDER     = settings.OUTPUT_FOLDER
SQL_FILE_PATH     = settings.SQL_FILE_PATH
ORDER_FILE_PREFIX = settings.ORDER_FILE_PREFIX


# ====================== 数据库操作 ======================
def connect_db():
    """建立数据库连接"""
    try:
        conn = pymysql.connect(**DB_CONFIG)
        print("✅ 数据库连接成功")
        return conn
    except Exception as e:
        print(f"❌ 数据库连接失败：{str(e)}")
        return None


def _get_dynamic_dates():
    """
    V3.2：运行时动态计算 SQL 日期范围
    优先读取 config/settings_override.json 中的扩展天数
    fallback 到 settings.py 默认值
    修复原因：settings.py 在进程启动时固化日期，override 文件的修改对当次运行无效
    """
    import json as _json
    from datetime import datetime, timedelta

    before = settings.SQL_EXTEND_DAYS_BEFORE
    after  = settings.SQL_EXTEND_DAYS_AFTER

    override_path = os.path.join("config", "settings_override.json")
    try:
        with open(override_path, "r", encoding="utf-8") as f:
            override = _json.load(f)
        before = override.get("SQL_EXTEND_DAYS_BEFORE", before)
        after  = override.get("SQL_EXTEND_DAYS_AFTER",  after)
    except Exception:
        pass

    today             = datetime.now()
    first_of_month    = today.replace(day=1)
    last_month_day    = first_of_month - timedelta(days=1)
    process_first_day = last_month_day.replace(day=1)
    process_last_day  = first_of_month - timedelta(days=1)

    start_date = (process_first_day - timedelta(days=before)).strftime("%Y-%m-%d 00:00:00")
    end_date   = (process_last_day  + timedelta(days=after)).strftime("%Y-%m-%d 23:59:59")
    return start_date, end_date


def load_sql_from_config():
    """
    从 config/SQL-config.txt 读取SQL模板
    V3.2：改为每次运行时动态计算日期范围，settings_override.json 的修改立即生效
    """
    try:
        with open(SQL_FILE_PATH, "r", encoding="utf-8") as f:
            sql = f.read().strip()

        start_date, end_date = _get_dynamic_dates()
        sql = sql.replace("{START_DATE}", start_date)
        sql = sql.replace("{END_DATE}",   end_date)

        print("✅ 成功读取外部SQL配置文件")
        print(f"   📅 数据捞取范围：{start_date} ~ {end_date}")
        return sql
    except Exception as e:
        print(f"❌ 读取SQL文件失败：{str(e)}")
        return None


def query_database(conn, sql):
    """执行SQL并返回DataFrame"""
    df = pd.read_sql(sql, conn)
    print(f"📊 查询完成，获取数据行数：{len(df)}")
    return df


# ====================== 导出并高速美化 ======================
def export_and_pretty(df):
    """导出Excel + 轻量高速美化"""
    ensure_folder(OUTPUT_FOLDER)
    month_str = get_last_month_str()          # 从utils统一获取，格式YYYYMM
    filename  = f"{ORDER_FILE_PREFIX}{month_str}.xlsx"
    save_path = os.path.join(OUTPUT_FOLDER, filename)

    with pd.ExcelWriter(save_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook(save_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 全表边框（使用utils统一样式）
    last_col   = chr(ord('A') + max_col - 1)
    data_range = f"A1:{last_col}{max_row}"
    for row in ws[data_range]:
        for cell in row:
            cell.border = FULL_BORDER

    # 表头样式（使用utils统一样式）
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = __import__('openpyxl').styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.alignment = CENTER_ALIGN

    # 日期列宽度与格式
    if max_col >= 2:
        ws.column_dimensions['B'].width = 20
        for cell in ws['B']:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # 运单号文本格式
    for cell in ws['A']:
        cell.number_format = "@"

    # 其余列自适应宽度
    for col_idx in range(1, max_col + 1):
        col_letter = chr(ord('A') + col_idx - 1)
        max_len = 0
        for cell in ws[col_letter][:100]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

    # 快递类型列宽度放最后，避免被自适应覆盖
    ws.column_dimensions['F'].width = 16

    wb.save(save_path)
    print(f"💾 导出并美化完成：{filename}")
    return save_path


# ====================== 主流程 ======================
def run_download_orders():
    """订单下载主流程"""
    conn = None
    try:
        conn = connect_db()
        if not conn:
            return None

        sql = load_sql_from_config()
        if not sql:
            return None

        df = query_database(conn, sql)
        if df.empty:
            print("⚠️ 查询结果为空，无需导出")
            return None

        export_and_pretty(df)
        return df

    except Exception as e:
        print(f"❌ 流程执行失败：{str(e)}")
        return None

    finally:
        if conn:
            conn.close()
            print("🔌 数据库连接已关闭")


if __name__ == "__main__":
    run_download_orders()
