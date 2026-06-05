# -*- coding: utf-8 -*-
import os
import json
import zipfile
import queue
import threading
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, Response, jsonify, send_file

from config import settings
from core import task_runner as _tr
from core import cache as _cache

express_bp = Blueprint("express", __name__)

EXPRESS_CONFIG_PATH = os.path.join("config", "express_config.json")
SETTINGS_OVERRIDE   = os.path.join("config", "settings_override.json")


# ====================== 工具函数 ======================
def get_process_month():
    today          = datetime.now()
    first_of_month = today.replace(day=1)
    last_month     = first_of_month - timedelta(days=1)
    return last_month.strftime("%Y-%m")


def get_prev_month(month_str):
    y, m = int(month_str[:4]), int(month_str[5:7])
    m -= 1
    if m == 0:
        m, y = 12, y - 1
    return f"{y:04d}-{m:02d}"


def load_express_config():
    try:
        with open(EXPRESS_CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"express_list": [
            {"name": "申通", "identify_column": "业务时间", "enabled": True},
            {"name": "中通", "identify_column": "扫描时间", "enabled": True}
        ]}


def load_settings_override():
    try:
        with open(SETTINGS_OVERRIDE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def download_month(month):
    output_folder = os.path.join("output", month)
    if not os.path.exists(output_folder):
        return "结果文件不存在", 404
    zip_path = os.path.join("output", f"{month}_结果.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(output_folder):
            for file in files:
                file_path = os.path.join(root, file)
                zf.write(file_path, os.path.relpath(file_path, "output"))
    return send_file(os.path.abspath(zip_path), as_attachment=True,
        download_name=f"{month}_对账结果.zip")


# ====================== 页面路由 ======================
@express_bp.route("/")
def dashboard():
    return render_template("dashboard.html",
        process_month=get_process_month(), active_page="dashboard")

@express_bp.route("/run")
def run_page():
    config = load_express_config()
    return render_template("run.html",
        process_month=get_process_month(), active_page="run",
        express_list=config.get("express_list", []))

@express_bp.route("/history")
def history_page():
    return render_template("history.html",
        process_month=get_process_month(), active_page="history")

@express_bp.route("/stats")
def stats_page():
    return render_template("stats.html",
        process_month=get_process_month(), active_page="stats")

@express_bp.route("/config")
def config_page():
    return render_template("config.html",
        process_month=get_process_month(), active_page="config")


# ====================== API：上传 ======================
@express_bp.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "没有收到文件"}), 400
    file          = request.files["file"]
    process_month = get_process_month()
    data_folder   = os.path.join("data", process_month)
    os.makedirs(data_folder, exist_ok=True)
    file.save(os.path.join(data_folder, file.filename))
    return jsonify({"ok": True, "msg": f"✅ {file.filename} 已上传到 data/{process_month}/"})


# ====================== API：触发运行 ======================
@express_bp.route("/run", methods=["POST"])
def run():
    with _tr.task_lock:
        if _tr.is_running:
            return jsonify({"ok": False, "msg": "⚠️ 当前有任务正在处理，请等待完成后再试"}), 429
        _tr.is_running = True

    while not _tr.log_queue.empty():
        try: _tr.log_queue.get_nowait()
        except queue.Empty: break

    process_month = get_process_month()
    data_folder   = os.path.join("data",   process_month)
    output_folder = os.path.join("output", process_month)

    xlsx_files = [
        f for f in os.listdir(data_folder)
        if f.endswith(".xlsx") and not f.startswith("~")
    ] if os.path.exists(data_folder) else []

    if not xlsx_files:
        _tr.is_running = False
        return jsonify({"ok": False, "msg": f"❌ data/{process_month}/ 目录为空，请先上传账单文件"}), 400

    threading.Thread(
        target=_tr.run_task,
        args=(data_folder, output_folder, process_month),
        daemon=True
    ).start()
    return jsonify({"ok": True, "msg": "任务已启动"})


# ====================== API：SSE日志 ======================
@express_bp.route("/logs")
def logs():
    def generate():
        while True:
            try:
                msg = _tr.log_queue.get(timeout=30)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg == "__DONE__":
                    break
            except queue.Empty:
                yield f"data: {json.dumps('__PING__', ensure_ascii=False)}\n\n"
    return Response(generate(), mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Access-Control-Allow-Origin": "*"})


# ====================== API：任务状态 ======================
@express_bp.route("/status")
def status():
    return jsonify({"running": _tr.is_running, "result": _tr.task_result})


# ====================== API：下载zip ======================
@express_bp.route("/download")
def download():
    return download_month(get_process_month())

@express_bp.route("/download/<month>")
def download_month_route(month):
    return download_month(month)


# ====================== API：历史记录 ======================
@express_bp.route("/history/api")
def history_api():
    output_base = "output"
    dir_mtime   = _cache._get_mtime(output_base)
    with _cache._cache_lock:
        if _cache._history_cache["result"] is not None and _cache._history_cache["dir_mtime"] == dir_mtime:
            return jsonify(_cache._history_cache["result"])
    if not os.path.exists(output_base):
        return jsonify([])
    records = []
    for folder in sorted(os.listdir(output_base), reverse=True):
        folder_path = os.path.join(output_base, folder)
        if not os.path.isdir(folder_path) or len(folder) != 7 or folder[4] != "-":
            continue
        log_path = os.path.join(folder_path, "run.log")
        record   = {"month": folder, "has_log": os.path.exists(log_path),
                    "run_count": 0, "last_result": "无记录", "last_time": "", "last_duration": ""}
        if os.path.exists(log_path):
            try:
                with open(log_path, "r", encoding="utf-8") as f:
                    content_log = f.read()
                record["run_count"] = content_log.count("【第")
                sections = content_log.split("【第")
                if len(sections) > 1:
                    last_section = sections[-1]
                    for line in last_section.split("\n"):
                        if "】" in line:
                            record["last_time"] = line.split("】")[-1].strip()
                            break
                    if "成功 ✅" in last_section: record["last_result"] = "成功"
                    elif "失败 ❌" in last_section: record["last_result"] = "失败"
                    for line in last_section.split("\n"):
                        if line.startswith("耗时："):
                            record["last_duration"] = line.replace("耗时：", "").strip()
                            break
            except Exception:
                record["last_result"] = "日志读取失败"
        records.append(record)
    with _cache._cache_lock:
        _cache._history_cache = {"dir_mtime": dir_mtime, "result": records}
    return jsonify(records)


# ====================== API：统计数据 ======================
@express_bp.route("/stats/<month>")
def stats(month):
    import pandas as pd
    output_base = "output"
    result_file = os.path.join(output_base, month, "最终对账结果.xlsx")
    file_mtime  = _cache._get_mtime(result_file)

    with _cache._cache_lock:
        cached = _cache._stats_cache.get(month)
        if cached and cached["file_mtime"] == file_mtime and file_mtime != -1:
            return jsonify(cached["result"])

    result = {
        "month":         month,
        "team_stats":    [],
        "monthly_trend": [],
        "team_summary":  [],
        "express_stats": [],
        "mom_change":    None,
        "anomalies":     [],
    }

    if os.path.exists(result_file):
        try:
            df = _cache._read_excel_cached(result_file)
            if df is not None:
                df_matched = df[df["所属团队"] != "未匹配"] if "所属团队" in df.columns else df
                if "所属团队" in df_matched.columns and "单票应付金额" in df_matched.columns:
                    df_matched = df_matched.copy()
                    df_matched["单票应付金额"] = pd.to_numeric(df_matched["单票应付金额"], errors="coerce").fillna(0)

                    team_group = df_matched.groupby("所属团队").agg(
                        amount=("单票应付金额", "sum"), count=("运单号", "count")
                    ).reset_index()
                    team_group = team_group[team_group["count"] > 0].sort_values("amount", ascending=False)
                    result["team_stats"] = [
                        {"team": row["所属团队"], "amount": round(float(row["amount"]), 2)}
                        for _, row in team_group.iterrows()
                    ]

                    summary = []
                    total_single = total_average = total_amount = 0
                    for team, group in df_matched.groupby("所属团队"):
                        single_amount = round(float(
                            group[group["实际计算方式"] == "单票"]["单票应付金额"].sum()
                        ), 2) if "实际计算方式" in group.columns else 0
                        average_count = len(group[group["实际计算方式"] == "全国均重"]) if "实际计算方式" in group.columns else 0
                        team_amount   = round(float(group["单票应付金额"].sum()), 2)
                        total_single += single_amount; total_average += average_count; total_amount += team_amount
                        summary.append({"team": team, "single_amount": single_amount,
                                        "average_count": average_count, "total_amount": team_amount})
                    summary.sort(key=lambda x: x["total_amount"], reverse=True)
                    summary.append({"team": "合计", "single_amount": round(total_single, 2),
                                    "average_count": total_average, "total_amount": round(total_amount, 2)})
                    result["team_summary"] = summary

                    if "快递类型" in df_matched.columns:
                        exp_group = df_matched.groupby("快递类型").agg(
                            amount=("单票应付金额", "sum"), count=("运单号", "count")
                        ).reset_index()
                        total_exp = float(exp_group["amount"].sum()) or 1
                        result["express_stats"] = [
                            {
                                "name":   str(row["快递类型"]),
                                "amount": round(float(row["amount"]), 2),
                                "count":  int(row["count"]),
                                "pct":    round(float(row["amount"]) / total_exp * 100, 1)
                            }
                            for _, row in exp_group.sort_values("amount", ascending=False).iterrows()
                        ]

                    df_full = df.copy()
                    if "结算重量" in df_full.columns:
                        df_full["结算重量"] = pd.to_numeric(df_full["结算重量"], errors="coerce").fillna(0)
                    if "单票应付金额" in df_full.columns:
                        df_full["单票应付金额"] = pd.to_numeric(df_full["单票应付金额"], errors="coerce")

                    anomaly_rules = [
                        ("重量异常", "high", lambda d: d[
                            d["结算重量"].notna() & ((d["结算重量"] <= 0) | (d["结算重量"] >= 50))
                        ] if "结算重量" in d.columns else d.iloc[0:0]),
                        ("省份为空", "mid", lambda d: d[
                            d["目的省份"].isna() | (d["目的省份"].astype(str).str.strip() == "")
                        ] if "目的省份" in d.columns else d.iloc[0:0]),
                        ("单票金额为零", "mid", lambda d: d[
                            (d["实际计算方式"] == "单票") &
                            (d["单票应付金额"].isna() | (d["单票应付金额"] == 0))
                        ] if "实际计算方式" in d.columns and "单票应付金额" in d.columns else d.iloc[0:0]),
                        ("未匹配团队", "mid", lambda d: d[
                            d["所属团队"] == "未匹配"
                        ] if "所属团队" in d.columns else d.iloc[0:0]),
                    ]

                    anomalies  = []
                    total_rows = len(df_full)
                    for atype, level, rule_fn in anomaly_rules:
                        try:
                            df_anom = rule_fn(df_full)
                            count   = len(df_anom)
                            if count == 0:
                                continue
                            samples = (
                                df_anom["运单号"].astype(str).head(5).tolist()
                                if "运单号" in df_anom.columns else []
                            )
                            pct = round(count / total_rows * 100, 1) if total_rows > 0 else 0
                            anomalies.append({"type": atype, "level": level,
                                              "count": count, "pct": pct, "samples": samples})
                        except Exception:
                            continue
                    anomalies.sort(key=lambda x: 0 if x["level"] == "high" else 1)
                    result["anomalies"] = anomalies
        except Exception as e:
            print(f"读取统计数据失败：{str(e)}")

    monthly = []
    if os.path.exists(output_base):
        for folder in sorted(os.listdir(output_base)):
            folder_path = os.path.join(output_base, folder)
            if not os.path.isdir(folder_path) or len(folder) != 7 or folder[4] != "-":
                continue
            f = os.path.join(folder_path, "最终对账结果.xlsx")
            if not os.path.exists(f):
                continue
            try:
                df_m = _cache._read_excel_cached(f, usecols=["单票应付金额"])
                if df_m is not None and "单票应付金额" in df_m.columns:
                    df_m["单票应付金额"] = pd.to_numeric(df_m["单票应付金额"], errors="coerce").fillna(0)
                    monthly.append({"month": folder, "total": round(float(df_m["单票应付金额"].sum()), 2)})
            except Exception:
                continue
    result["monthly_trend"] = monthly

    prev_month    = get_prev_month(month)
    current_total = next((m["total"] for m in monthly if m["month"] == month), None)
    prev_total    = next((m["total"] for m in monthly if m["month"] == prev_month), None)
    if current_total is not None and prev_total is not None and prev_total != 0:
        change     = round(current_total - prev_total, 2)
        change_pct = round(change / prev_total * 100, 1)
        result["mom_change"] = {
            "prev_month": prev_month, "prev_total": prev_total,
            "change": change, "change_pct": change_pct
        }

    with _cache._cache_lock:
        _cache._stats_cache[month] = {"file_mtime": file_mtime, "result": result}
    return jsonify(result)


# ====================== API：未匹配分析 ======================
@express_bp.route("/unmatched/<month>")
def unmatched(month):
    import pandas as pd
    result_file = os.path.join("output", month, "最终对账结果.xlsx")
    if not os.path.exists(result_file):
        return jsonify({"ok": False, "msg": f"{month} 对账结果文件不存在"}), 404
    file_mtime = _cache._get_mtime(result_file)
    with _cache._cache_lock:
        cached = _cache._unmatched_cache.get(month)
        if cached and cached["file_mtime"] == file_mtime:
            return jsonify(cached["result"])
    df = _cache._read_excel_cached(result_file)
    if df is None:
        return jsonify({"ok": False, "msg": "读取文件失败"}), 500
    total           = len(df)
    df_unmatched    = df[df["所属团队"] == "未匹配"] if "所属团队" in df.columns else df
    unmatched_count = len(df_unmatched)
    matched_count   = total - unmatched_count
    ratio           = round(unmatched_count / total * 100, 1) if total > 0 else 0
    by_express      = {}
    if "快递类型" in df_unmatched.columns:
        for t, g in df_unmatched.groupby("快递类型"):
            by_express[str(t)] = len(g)
    samples = df_unmatched["运单号"].astype(str).head(5).tolist() if "运单号" in df_unmatched.columns else []
    result  = {"ok": True, "month": month, "total": total, "matched": matched_count,
               "unmatched": unmatched_count, "ratio": ratio, "by_express": by_express, "samples": samples}
    with _cache._cache_lock:
        _cache._unmatched_cache[month] = {"file_mtime": file_mtime, "result": result}
    return jsonify(result)


# ====================== API：异常运单下载 ======================
@express_bp.route("/anomaly/download/<month>")
def anomaly_download(month):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    import tempfile

    result_file = os.path.join("output", month, "最终对账结果.xlsx")
    if not os.path.exists(result_file):
        return jsonify({"ok": False, "msg": f"{month} 对账结果文件不存在"}), 404

    df = _cache._read_excel_cached(result_file)
    if df is None:
        return jsonify({"ok": False, "msg": "读取文件失败"}), 500

    df = df.copy()
    if "结算重量" in df.columns:
        df["结算重量"] = pd.to_numeric(df["结算重量"], errors="coerce").fillna(0)
    if "单票应付金额" in df.columns:
        df["单票应付金额"] = pd.to_numeric(df["单票应付金额"], errors="coerce")

    anomaly_map = {}

    def mark(mask, atype, reason):
        for idx in df[mask].index:
            if idx not in anomaly_map:
                anomaly_map[idx] = {"types": [], "reasons": []}
            anomaly_map[idx]["types"].append(atype)
            anomaly_map[idx]["reasons"].append(reason)

    if "结算重量" in df.columns:
        mark(df["结算重量"] <= 0, "重量异常", "结算重量为0或负数，可能录入错误或账单格式问题，建议核对原始账单")
        mark(df["结算重量"] >= 50, "重量异常", "结算重量≥50kg，超出正常范围，建议人工核实是否为实际重量")
    if "目的省份" in df.columns:
        mask_prov = df["目的省份"].isna() | (df["目的省份"].astype(str).str.strip() == "")
        mark(mask_prov, "省份为空", "目的省份字段为空，计费模式可能判断不准确，建议检查原始账单格式")
    if "实际计算方式" in df.columns and "单票应付金额" in df.columns:
        mask_zero = (df["实际计算方式"] == "单票") & \
                    (df["单票应付金额"].isna() | (df["单票应付金额"] == 0))
        mark(mask_zero, "单票金额为零", "计费方式为单票但应付金额为0，可能报价表缺少该省份数据，建议检查申通/中通报价配置")
    if "所属团队" in df.columns:
        mark(df["所属团队"] == "未匹配", "未匹配团队",
             "运单号在数据库订单中未找到对应团队，可能原因：①运单不属于本店 ②SQL日期范围未覆盖 ③运单号格式不一致")

    if not anomaly_map:
        return jsonify({"ok": False, "msg": f"{month} 未发现任何异常运单"}), 404

    anom_indices = list(anomaly_map.keys())
    df_anom = df.loc[anom_indices].copy()
    df_anom["异常类型"]     = ["/".join(anomaly_map[i]["types"])   for i in anom_indices]
    df_anom["异常原因说明"] = ["\n".join(anomaly_map[i]["reasons"]) for i in anom_indices]

    level_order = {"重量异常": 0, "省份为空": 1, "单票金额为零": 2, "未匹配团队": 3}
    df_anom["_sort"] = df_anom["异常类型"].apply(
        lambda x: min(level_order.get(t, 9) for t in x.split("/"))
    )
    df_anom = df_anom.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    df_anom.to_excel(tmp.name, index=False, engine="openpyxl")

    wb = load_workbook(tmp.name)
    ws = wb.active
    headers   = [cell.value for cell in ws[1]]
    type_col   = headers.index("异常类型")     + 1 if "异常类型"     in headers else None
    reason_col = headers.index("异常原因说明") + 1 if "异常原因说明" in headers else None

    fill_high = PatternFill(start_color="4D1010", end_color="4D1010", fill_type="solid")
    fill_mid  = PatternFill(start_color="4D3010", end_color="4D3010", fill_type="solid")
    fill_head = PatternFill(start_color="1E2235", end_color="1E2235", fill_type="solid")
    high_types = {"重量异常"}

    for cell in ws[1]:
        cell.fill      = fill_head
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if type_col:
            tc    = row[type_col - 1]
            types = str(tc.value or "").split("/")
            tc.fill = fill_high if any(t in high_types for t in types) else fill_mid
            tc.font = Font(bold=True, color="FFFFFF", size=11)
            tc.alignment = Alignment(horizontal="center", vertical="center")
        if reason_col:
            row[reason_col - 1].alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = {}
    for col_idx, header in enumerate(headers, 1):
        if header == "运单号":          col_widths[col_idx] = 22
        elif header == "异常类型":      col_widths[col_idx] = 18
        elif header == "异常原因说明":  col_widths[col_idx] = 55
        elif header in ("目的省份", "目的城市", "快递类型", "所属团队"): col_widths[col_idx] = 14
        else:                           col_widths[col_idx] = 12
    for ci, w in col_widths.items():
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(tmp.name)

    return send_file(tmp.name, as_attachment=True,
        download_name=f"{month}_异常运单_{len(df_anom)}条.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ====================== API：对账结果预览 ======================
@express_bp.route("/preview/<month>")
def preview(month):
    import pandas as pd
    result_file = os.path.join("output", month, "最终对账结果.xlsx")
    if not os.path.exists(result_file):
        return jsonify({"ok": False, "msg": f"{month} 对账结果文件不存在"}), 404
    df = _cache._read_excel_cached(result_file)
    if df is None:
        return jsonify({"ok": False, "msg": "读取文件失败"}), 500
    df      = df.fillna("")
    page    = int(request.args.get("page", 1))
    size    = int(request.args.get("size", 100))
    filter_ = request.args.get("filter", "all")
    keyword = request.args.get("keyword", "").strip()
    total           = len(df)
    matched_count   = len(df[df["所属团队"] != "未匹配"]) if "所属团队" in df.columns else 0
    unmatched_count = total - matched_count
    if "所属团队" in df.columns and "实际计算方式" in df.columns:
        if filter_ == "matched":     df = df[df["所属团队"] != "未匹配"]
        elif filter_ == "unmatched": df = df[df["所属团队"] == "未匹配"]
        elif filter_ == "single":    df = df[df["实际计算方式"] == "单票"]
        elif filter_ == "average":   df = df[df["实际计算方式"] == "全国均重"]
    if keyword:
        mask = pd.Series([False] * len(df), index=df.index)
        for col in ["运单号", "所属团队"]:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.contains(keyword, na=False)
        df = df[mask]
    filtered_count = len(df)
    page_df        = df.iloc[(page - 1) * size: page * size]
    show_cols      = [c for c in ["运单号", "所属团队", "目的省份", "结算重量", "快递类型", "实际计算方式", "单票应付金额"] if c in page_df.columns]
    return jsonify({"ok": True, "month": month, "total": total, "matched": matched_count,
        "unmatched": unmatched_count, "filtered": filtered_count, "page": page, "size": size,
        "total_pages": max(1, -(-filtered_count // size)), "rows": page_df[show_cols].to_dict(orient="records")})


# ====================== API：价格配置 ======================
@express_bp.route("/price/get")
def price_get():
    import pandas as pd
    price_file = os.path.join("config", "price_config.xlsx")
    if not os.path.exists(price_file):
        return jsonify({"ok": False, "msg": "price_config.xlsx 不存在"}), 404
    result = {"ok": True, "shentong": [], "zhongtong": [], "charge": []}
    try:
        for sheet, key in [("申通报价", "shentong"), ("中通报价", "zhongtong")]:
            df = pd.read_excel(price_file, sheet_name=sheet)
            for _, row in df.iterrows():
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "": continue
                result[key].append({
                    "province":    str(row.iloc[0]).strip(),
                    "fee_3kg":     float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0,
                    "fee_over3kg": float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0,
                    "unit_price":  float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
                })
        df_charge = pd.read_excel(price_file, sheet_name="客户快递加收单价信息记录")
        for _, row in df_charge.iterrows():
            if pd.isna(row.iloc[10]) or pd.isna(row.iloc[11]): continue
            type_name = str(row.iloc[10]).strip()
            if type_name in ["申通", "中通"]:
                result["charge"].append({"type": type_name, "price": float(row.iloc[11])})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500
    return jsonify(result)


@express_bp.route("/price/save", methods=["POST"])
def price_save():
    from openpyxl import load_workbook
    data       = request.json
    price_file = os.path.join("config", "price_config.xlsx")
    if not os.path.exists(price_file):
        return jsonify({"ok": False, "msg": "price_config.xlsx 不存在"}), 404
    try:
        wb = load_workbook(price_file)
        for sheet_key, sheet_name in [("shentong", "申通报价"), ("zhongtong", "中通报价")]:
            if sheet_key not in data or sheet_name not in wb.sheetnames:
                continue
            ws       = wb[sheet_name]
            new_rows = data[sheet_key]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            for i, item in enumerate(new_rows, start=2):
                ws.cell(row=i, column=1).value = item["province"]
                ws.cell(row=i, column=2).value = item["fee_3kg"]
                ws.cell(row=i, column=4).value = item["fee_over3kg"]
                ws.cell(row=i, column=5).value = item["unit_price"]
        if "charge" in data and "客户快递加收单价信息记录" in wb.sheetnames:
            ws = wb["客户快递加收单价信息记录"]
            for row in ws.iter_rows(min_row=2):
                if row[10].value and str(row[10].value).strip() in ["申通", "中通"]:
                    for item in data["charge"]:
                        if item["type"] == str(row[10].value).strip():
                            row[11].value = item["price"]
        wb.save(price_file)
        return jsonify({"ok": True, "msg": "价格配置已保存"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ====================== API：快递配置 ======================
@express_bp.route("/express/config")
def express_config():
    return jsonify({"ok": True, **load_express_config()})

@express_bp.route("/express/save", methods=["POST"])
def express_save():
    try:
        with open(EXPRESS_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(request.json, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True, "msg": "快递配置已保存"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ====================== API：运行参数配置 ======================
@express_bp.route("/settings/get")
def settings_get():
    override = load_settings_override()
    return jsonify({
        "ok":                 True,
        "extend_days_before": override.get("SQL_EXTEND_DAYS_BEFORE", settings.SQL_EXTEND_DAYS_BEFORE),
        "extend_days_after":  override.get("SQL_EXTEND_DAYS_AFTER",  settings.SQL_EXTEND_DAYS_AFTER),
        "process_month":      settings.PROCESS_MONTH,
        "sql_start_date":     settings.SQL_START_DATE,
        "sql_end_date":       settings.SQL_END_DATE,
    })


@express_bp.route("/settings/save", methods=["POST"])
def settings_save():
    data = request.json
    try:
        override = load_settings_override()
        before = int(data.get("extend_days_before", override.get("SQL_EXTEND_DAYS_BEFORE", 15)))
        after  = int(data.get("extend_days_after",  override.get("SQL_EXTEND_DAYS_AFTER",  5)))
        if before < 0 or after < 0:
            return jsonify({"ok": False, "msg": "天数不能为负数"}), 400
        override["SQL_EXTEND_DAYS_BEFORE"] = before
        override["SQL_EXTEND_DAYS_AFTER"]  = after
        os.makedirs("config", exist_ok=True)
        with open(SETTINGS_OVERRIDE, "w", encoding="utf-8") as f:
            json.dump(override, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True, "msg": f"已保存：前扩展 {before} 天，后扩展 {after} 天"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500
