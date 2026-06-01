# -*- coding: utf-8 -*-
"""
数据库订单下载模块（V2.1 稳定版）
==================================================
【框架说明】
  功能：从阿里云MySQL下载订单数据
       → 读取外部SQL语句
       → 导出Excel并高速美化
       → 保存到 output/YYYY-MM/毅播快递数据_YYYYMM.xlsx
  结构：配置 → 工具 → 数据库连接 → 查询 → 导出美化

【V2.1 变更】
  - 输出路径跟随 settings.OUTPUT_FOLDER（含月份子目录）
  - 其余逻辑完全保留原V1.4框架

【可独立运行】
==================================================
"""
import pymysql
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 【仅修改这一行】
from config import settings

# ====================== 配置（从settings读取）======================
DB_CONFIG         = settings.DB_CONFIG
OUTPUT_FOLDER     = settings.OUTPUT_FOLDER      # 已含月份子目录：output/YYYY-MM
SQL_FILE_PATH     = settings.SQL_FILE_PATH
ORDER_FILE_PREFIX = settings.ORDER_FILE_PREFIX

# ====================== 工具函数 ======================
def ensure_folder(folder_path):
    """确保文件夹存在"""
    os.makedirs(folder_path, exist_ok=True)

def get_last_month_str():
    """获取上个月：YYYYMM格式"""
    today = datetime.now()
    first_day_this_month = datetime(today.year, today.month, 1)
    last_day_last_month  = first_day_this_month - timedelta(days=1)
    return last_day_last_month.strftime("%Y%m")

# ====================== 数据库操作 ======================
def connect_db():
    """建立数据库连接"""
    try:
        conn = pymysql.connect(**DB_CONFIG)
        print("✅ 数据库连接成功")
        return conn
    except Exception as e:
        print(f"❌ 数据库连接失败：{str(e)}")
        return None

def load_sql_from_config():
    """
    从 config/SQL-config.txt 读取SQL模板
    并自动将 {START_DATE} / {END_DATE} 替换为本次处理月份的动态日期范围
    业务逻辑：日期范围由 settings.py 自动计算，无需手动修改 SQL 文件
    """
    try:
        with open(SQL_FILE_PATH, "r", encoding="utf-8") as f:
            sql = f.read().strip()

        # 注入动态日期范围
        sql = sql.replace("{START_DATE}", settings.SQL_START_DATE)
        sql = sql.replace("{END_DATE}",   settings.SQL_END_DATE)

        print("✅ 成功读取外部SQL配置文件")
        print(f"   📅 数据捞取范围：{settings.SQL_START_DATE} ~ {settings.SQL_END_DATE}")
        return sql
    except Exception as e:
        print(f"❌ 读取SQL文件失败：{str(e)}")
        return None

def query_database(conn, sql):
    """执行SQL并返回DataFrame"""
    df = pd.read_sql(sql, conn)
    print(f"📊 查询完成，获取数据行数：{len(df)}")
    return df

# ====================== 导出并高速美化 ======================
def export_and_pretty(df):
    """导出Excel + 轻量高速美化"""
    ensure_folder(OUTPUT_FOLDER)
    month_str = get_last_month_str()
    filename  = f"{ORDER_FILE_PREFIX}{month_str}.xlsx"
    save_path = os.path.join(OUTPUT_FOLDER, filename)

    # 快速导出
    with pd.ExcelWriter(save_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook(save_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 样式定义
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font  = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 全表边框
    last_col   = chr(ord('A') + max_col - 1)
    data_range = f"A1:{last_col}{max_row}"
    for row in ws[data_range]:
        for cell in row:
            cell.border = border

    # 表头样式
    for cell in ws[1]:
        cell.font      = header_font
        cell.alignment = center_align

    # 日期列宽度与格式
    if max_col >= 2:
        ws.column_dimensions['B'].width = 20
        for cell in ws['B']:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # 运单号文本格式
    for cell in ws['A']:
        cell.number_format = "@"

    # 其余列自适应宽度
    for col_idx in range(1, max_col + 1):
        col_letter = chr(ord('A') + col_idx - 1)
        max_len = 0
        for cell in ws[col_letter][:100]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

    # 【关键】快递类型列宽度放最后，避免被自适应覆盖
    ws.column_dimensions['F'].width = 16

    wb.save(save_path)
    print(f"💾 导出并美化完成：{filename}")
    return save_path

# ====================== 主流程 ======================
def run_download_orders():
    """订单下载主流程"""
    conn = None
    try:
        conn = connect_db()
        if not conn:
            return None

        sql = load_sql_from_config()
        if not sql:
            return None

        df = query_database(conn, sql)
        if df.empty:
            print("⚠️ 查询结果为空，无需导出")
            return None

        export_and_pretty(df)
        return df

    except Exception as e:
        print(f"❌ 流程执行失败：{str(e)}")
        return None

    finally:
        if conn:
            conn.close()
            print("🔌 数据库连接已关闭")

# 独立运行入口
if __name__ == "__main__":
    run_download_orders()