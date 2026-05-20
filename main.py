# -*- coding: utf-8 -*-
"""
快递对账自动化系统
最终修复版：彻底删除脏数据+全表边框
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ====================== 读取Excel文件 ======================
def read_excel(file_path):
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        return None
    try:
        df = pd.read_excel(file_path, engine="openpyxl")
        print(f"✅ 读取成功：{os.path.basename(file_path)} | 行数：{len(df)}")
        return df
    except Exception as e:
        print(f"❌ 读取失败：{str(e)}")
        return None

# ====================== 深度清洗（双重过滤，干掉脏数据） ======================
def deep_clean_df(df):
    if df is None or df.empty:
        return None

    # 1. 删除全空行
    df = df.dropna(how="all")
    # 2. 删除重复行
    df = df.drop_duplicates()
    # 3. 重置索引
    df = df.reset_index(drop=True)
    # 4. 关键：只保留至少有2个非空值的行（干掉只有一个单元格有值的脏数据）
    df = df[df.count(axis=1) > 1]
    # 5. 重置索引
    df = df.reset_index(drop=True)
    return df

# ====================== 处理申通账单 ======================
def handle_shentong_df(df_st):
    df_clean = deep_clean_df(df_st)
    need_cols = ["业务时间", "运单号", "订单目的省份", "订单目的城市", "结算重量"]
    df_use = df_clean[need_cols].copy()
    df_use.rename(columns={
        "订单目的省份": "目的省份",
        "订单目的城市": "目的城市"
    }, inplace=True)
    df_use["运单号"] = df_use["运单号"].astype(str)
    df_use["快递"] = "申通"
    df_use["团队"] = ""
    return df_use

# ====================== 处理中通账单 ======================
def handle_zhongtong_df(df_zt):
    df_clean = deep_clean_df(df_zt)
    need_cols = ["扫描时间", "运单号", "目的地", "目的地市", "结算重量"]
    df_use = df_clean[need_cols].copy()
    df_use.rename(columns={
        "扫描时间": "业务时间",
        "目的地": "目的省份",
        "目的地市": "目的城市"
    }, inplace=True)
    df_use["运单号"] = df_use["运单号"].astype(str)
    df_use["快递"] = "中通"
    df_use["团队"] = ""
    return df_use

# ====================== 合并表格 + 最终清洗 ======================
def merge_two_df(df_st_final, df_zt_final):
    final_cols = ["业务时间", "运单号", "目的省份", "目的城市", "结算重量", "快递", "团队"]
    df_total = pd.concat([df_st_final, df_zt_final], ignore_index=True)
    df_total = df_total[final_cols]
    # 合并后再执行一次深度清洗，确保脏数据全部删除
    df_total = deep_clean_df(df_total)
    # 额外保险：强制删除运单号为空的行
    df_total = df_total.dropna(subset=["运单号"])
    df_total = df_total.reset_index(drop=True)
    print(f"\n🔗 合并完成，总数据：{len(df_total)} 条")
    return df_total

# ====================== Excel格式美化（全表边框） ======================
def save_and_style_excel(df_total, output_path):
    df_total.to_excel(output_path, index=False, engine="openpyxl")
    wb = load_workbook(output_path)
    ws = wb.active

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表头：加粗 + 居中 + 边框
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 所有数据区域加边框（含团队列）
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = border

    # 运单号 = 文本格式
    for cell in ws["B"]:
        if cell.value is not None:
            cell.number_format = "@"

    # 结算重量 = 数字格式
    for cell in ws["E"]:
        if cell.value is not None:
            cell.number_format = "0.00"

    # 固定列宽
    widths = {"A":22, "B":25, "C":15, "D":15, "E":12, "F":10, "G":10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"💾 美化完成：{output_path}")

# ====================== 主流程 ======================
def main():
    print("=" * 60)
    print("        中通+申通快递对账合并工具（脏数据彻底修复版）")
    print("=" * 60)

    path_st = "data/2026年4月钟村毅播云仓对帐单.xlsx"
    path_zt = "data/咸辣甜服饰-4月份.xlsx"
    df_st_raw = read_excel(path_st)
    df_zt_raw = read_excel(path_zt)

    df_st_done = handle_shentong_df(df_st_raw)
    df_zt_done = handle_zhongtong_df(df_zt_raw)

    df_all = merge_two_df(df_st_done, df_zt_done)

    save_and_style_excel(df_all, "data/清洗合并总账单.xlsx")

    print("\n🎉 全部流程执行完毕！脏数据已彻底删除！")

if __name__ == "__main__":
    main()