# -*- coding: utf-8 -*-
"""
客户对账账单拆分 V3.3
==================================================
【核心功能】
1. 按团队拆分账单，自动生成标准化Excel格式
2. 运费核算：全国均重/单票/新西1-3kg/新西1kg内加收费用计算
3. 专项统计：新西1kg订单数、1%占比判断、应结算单数核算
4. 文件命名：{合计金额}-{团队名}_{业务年月}_快递加收费.xlsx
5. 0费用过滤：普通团队合计金额为0不生成文件

【V2.4 架构整理】
  - 删除重复的 init_folder 函数，改从 utils 导入 ensure_folder
  - 样式常量改从 utils 导入，保持全项目Excel风格一致
  - 所有计费逻辑、千耀传媒专项、统计表完全保留V2.0框架

【可独立运行】
==================================================
"""
import os
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from config import settings
from utils import (
    ensure_folder, FULL_BORDER, HEADER_FONT, BOLD_FONT,
    CENTER_ALIGN, LEFT_ALIGN, RIGHT_ALIGN, WRAP_CENTER_ALIGN
)

# ===================== 全局配置 =====================
SOURCE_FILE = os.path.join(settings.OUTPUT_FOLDER, settings.RESULT_FILE)
CONFIG_FILE = os.path.join(settings.CONFIG_FOLDER, "price_config.xlsx")
SAVE_DIR    = os.path.join(settings.OUTPUT_FOLDER, "客户账单")

GROUP_COL      = "所属团队"
FILTER_INVALID = True
TEST_TEAMS     = []
SPECIAL_TEAM   = "千耀传媒"

# ===================== 补充样式（split_bill专用）=====================
RED_BOLD_FONT = Font(bold=True, color="FF0000")

COL_WIDTH = {
    'A': 20, 'B': 18, 'C': 14, 'D': 12, 'E': 8,
    'F': 12, 'G': 12, 'H': 13, 'I': 13,
    'L': 8,  'M': 18, 'N': 12, 'O': 14,
    'P': 14, 'Q': 12, 'R': 15, 'S': 15
}
ROW_HEIGHT  = 24
TEAM_CONFIG = {}


# ===================== 工具函数 =====================
def safe_file_name(raw_name: str) -> str:
    illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    name = str(raw_name)
    for char in illegal_chars:
        name = name.replace(char, "_")
    return name


def load_team_config():
    global TEAM_CONFIG
    if not os.path.exists(CONFIG_FILE):
        print(f"【警告】配置文件不存在 {CONFIG_FILE}")
        return
    try:
        df = pd.read_excel(CONFIG_FILE, engine="openpyxl", sheet_name="客户快递加收单价信息记录")
        for _, row in df.iterrows():
            team = str(row.iloc[3]).strip()
            if not team or team == "nan":
                continue
            st_avg = float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0.0
            zt_avg = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0.0
            st_fee = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0.0
            zt_fee = float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0.0
            TEAM_CONFIG[team] = (st_avg, zt_avg, st_fee, zt_fee)
        print(f"【配置】加载 {len(TEAM_CONFIG)} 个团队规则")
    except Exception as e:
        print(f"【错误】读取配置失败: {str(e)}")


def get_team_rule(team_name: str) -> tuple:
    return TEAM_CONFIG.get(team_name, (0.0, 0.0, 0.0, 0.0))


def set_worksheet_style(ws, max_row: int, max_col: int):
    for col, width in COL_WIDTH.items():
        ws.column_dimensions[col].width = width
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = ROW_HEIGHT
    for row in range(1, max_row + 1):
        ws[f"B{row}"].number_format = "@"
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border    = FULL_BORDER
            cell.alignment = CENTER_ALIGN
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


# ===================== 核心汇总统计（完整保留V2.0逻辑）=====================
def add_summary_area(ws, team_df: pd.DataFrame, team_name: str) -> float:
    st_avg_all, zt_avg_all, st_fee_all, zt_fee_all = get_team_rule(team_name)

    col_start = 12
    col_end   = 19

    total_order  = len(team_df)
    xixi_order   = len(team_df[(team_df["结算重量"] < 1) & (team_df["目的省份"].str.contains("新疆|西藏", na=False))])
    threshold    = total_order * 0.01
    is_over_flag = "是" if xixi_order >= threshold else "否"
    settle_order = math.ceil(xixi_order - threshold) if is_over_flag == "是" else 0
    r11_value    = settle_order * 10

    # 汇总大标题
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
    title_cell           = ws.cell(row=1, column=col_start)
    title_cell.value     = "加收费汇总"
    title_cell.font      = HEADER_FONT
    title_cell.fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = CENTER_ALIGN
    title_cell.border    = FULL_BORDER

    if team_name == SPECIAL_TEAM:
        summary_headers = ["序号","实际计算方式","快递类型","发货单量","结算重量","平均重量","超出重量","应付金额"]
        calc_list = [
            ("全国均重","申通"), ("全国均重","中通"),
            ("单票","申通"), ("新西1-3公斤","申通"),
            ("新西1kg内（包1%）","中通"), ("合计","")
        ]
    else:
        summary_headers = ["序号","实际计算方式","发货单量","结算重量","平均重量","超出重量","应付金额"]
        calc_list = [
            ("全国均重",""), ("单票",""), ("新西1-3公斤",""),
            ("新西1kg内（包1%）",""), ("合计","")
        ]

    for idx, text in enumerate(summary_headers):
        cell           = ws.cell(row=2, column=col_start + idx)
        cell.value     = text
        cell.border    = FULL_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font      = BOLD_FONT

    summary_rows  = []
    total_fee_all = 0.0

    if team_name == SPECIAL_TEAM:
        for idx, (calc_name, express_type) in enumerate(calc_list):
            seq = idx + 1
            order_cnt = total_weight = avg_weight = exceed_weight = row_fee = 0.0

            if calc_name != "合计":
                filter_df = team_df[(team_df["实际计算方式"] == calc_name) & (team_df["快递类型"] == express_type)]
                order_cnt = len(filter_df)
                if order_cnt > 0:
                    total_weight = round(filter_df["结算重量"].sum(), 2)
                    avg_weight   = round(total_weight / order_cnt, 2)

            use_avg = st_avg_all if express_type == "申通" else zt_avg_all
            use_fee = st_fee_all if express_type == "申通" else zt_fee_all

            if calc_name == "全国均重":
                exceed_weight  = max(round(avg_weight - use_avg, 2), 0.0)
                single_fee     = round((exceed_weight / 0.1) * use_fee, 2)
                row_fee        = round(single_fee * order_cnt, 2)
                total_fee_all += row_fee
            elif calc_name in ("单票", "新西1-3公斤"):
                _w = pd.to_numeric(filter_df["结算重量"], errors="coerce").dropna()
                fee_sum        = ((_w - use_avg).clip(lower=0) / 0.1 * use_fee).round(2).sum()
                row_fee        = round(fee_sum, 2)
                total_fee_all += row_fee
            elif calc_name == "新西1kg内（包1%）":
                row_fee        = r11_value
                total_fee_all += row_fee

            if calc_name == "新西1kg内（包1%）":
                row_data = [seq, calc_name, express_type, "", "", "", "", row_fee]
            elif calc_name == "合计":
                row_data = [seq, calc_name, "", total_order, "", "", "", "=SUM(S3:S7)"]
            else:
                row_data = [seq, calc_name, express_type, order_cnt, total_weight, avg_weight, exceed_weight, row_fee]

            if calc_name in ("单票", "新西1-3公斤", "新西1kg内（包1%）", "合计"):
                row_data[4] = row_data[5] = row_data[6] = ""

            summary_rows.append(row_data)
    else:
        for idx, (calc_name, _) in enumerate(calc_list):
            seq = idx + 1
            order_cnt = total_weight = avg_weight = exceed_weight = row_fee = 0.0

            if calc_name != "合计":
                filter_df = team_df[team_df["实际计算方式"] == calc_name]
                order_cnt = len(filter_df)
                if order_cnt > 0:
                    total_weight = round(filter_df["结算重量"].sum(), 2)
                    avg_weight   = round(total_weight / order_cnt, 2)

            if calc_name == "全国均重":
                exceed_weight  = max(round(avg_weight - st_avg_all, 2), 0.0)
                single_fee     = round((exceed_weight / 0.1) * st_fee_all, 2)
                row_fee        = round(single_fee * order_cnt, 2)
                total_fee_all += row_fee
            elif calc_name in ("单票", "新西1-3公斤"):
                _w = pd.to_numeric(filter_df["结算重量"], errors="coerce").dropna()
                fee_sum        = ((_w - st_avg_all).clip(lower=0) / 0.1 * st_fee_all).round(2).sum()
                row_fee        = round(fee_sum, 2)
                total_fee_all += row_fee
            elif calc_name == "新西1kg内（包1%）":
                row_fee        = r11_value
                total_fee_all += row_fee

            if calc_name == "新西1kg内（包1%）":
                row_data = [seq, calc_name, "", "", "", "", row_fee]
            elif calc_name == "合计":
                row_data = [seq, calc_name, total_order, "", "", "", "=SUM(R3:R6)"]
            else:
                row_data = [seq, calc_name, order_cnt, total_weight, avg_weight, exceed_weight, row_fee]

            if calc_name in ("单票", "新西1-3公斤", "新西1kg内（包1%）", "合计"):
                row_data[3] = row_data[4] = row_data[5] = ""

            summary_rows.append(row_data)

    total_fee_all = round(total_fee_all, 2)

    for row_idx, row_data in enumerate(summary_rows, start=3):
        for col_idx, val in enumerate(row_data):
            cell        = ws.cell(row=row_idx, column=col_start + col_idx)
            cell.value  = val
            cell.border = FULL_BORDER
            if col_idx == 0:
                cell.alignment = CENTER_ALIGN
            elif col_idx == 1:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = RIGHT_ALIGN
            if row_data[1] == "合计":
                if col_idx == 1:
                    cell.font      = BOLD_FONT
                    cell.alignment = CENTER_ALIGN
                if (team_name == SPECIAL_TEAM and col_idx == 7) or (team_name != SPECIAL_TEAM and col_idx == 6):
                    cell.font = RED_BOLD_FONT
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    # 专项统计表
    if team_name == SPECIAL_TEAM:
        stat_row_t1, stat_row_t2, stat_row_data = 10, 11, 12
        stat_headers = ["序号","实际计算方式","快递类型","总发货单量","新西1kg内订单数","是否超1%","新西1kg内应结算单数","新西1kg内应付金额"]
    else:
        stat_row_t1, stat_row_t2, stat_row_data = 9, 10, 11
        stat_headers = ["序号","实际计算方式","总发货单量","新西1kg内订单数","是否超1%","新西1kg内应结算单数","新西1kg内应付金额"]

    for idx in range(len(stat_headers)):
        col = col_start + idx
        ws.merge_cells(start_row=stat_row_t1, start_column=col, end_row=stat_row_t2, end_column=col)
        cell           = ws.cell(row=stat_row_t1, column=col)
        cell.value     = stat_headers[idx]
        cell.alignment = WRAP_CENTER_ALIGN
        cell.font      = BOLD_FONT
        cell.border    = FULL_BORDER

    if team_name == SPECIAL_TEAM:
        cell           = ws.cell(row=stat_row_data, column=col_start+2, value="中通")
        cell.alignment = CENTER_ALIGN
        cell.border    = FULL_BORDER

    stat_data = ["1", "新西1kg内（包1%）", total_order, xixi_order, is_over_flag, settle_order, r11_value]
    if team_name == SPECIAL_TEAM:
        stat_data.insert(2, "")

    for idx, val in enumerate(stat_data):
        cell        = ws.cell(row=stat_row_data, column=col_start + idx)
        cell.value  = val
        cell.border = FULL_BORDER
        if idx == 0:
            cell.alignment = CENTER_ALIGN
        elif idx == 1:
            cell.alignment = LEFT_ALIGN
        else:
            cell.alignment = RIGHT_ALIGN

    for row in range(stat_row_t1, stat_row_data + 1):
        for col in range(col_start, col_start + len(stat_headers)):
            ws.cell(row=row, column=col).border = FULL_BORDER

    for r in [stat_row_t1, stat_row_t2, stat_row_data]:
        ws.row_dimensions[r].height = ROW_HEIGHT

    return total_fee_all


# ===================== 数据读取与文件拆分 =====================
def load_source_data() -> pd.DataFrame:
    if not os.path.exists(SOURCE_FILE):
        raise FileNotFoundError(f"原始数据文件不存在 {SOURCE_FILE}")
    df = pd.read_excel(SOURCE_FILE, engine="openpyxl")
    print(f"【数据】原始条数：{len(df)}")
    if FILTER_INVALID:
        df = df[(df[GROUP_COL].notna()) & (df[GROUP_COL] != "未匹配")]
        print(f"【数据】有效条数：{len(df)}")
    df["运单号"] = df["运单号"].astype(str)
    return df


def split_by_group(df: pd.DataFrame):
    group_list = df.groupby(GROUP_COL)
    print(f"\n【分组】共识别 {len(group_list)} 个团队")
    print("===== 开始拆分账单 =====")

    for team_name, team_data in group_list:
        if TEST_TEAMS and team_name not in TEST_TEAMS:
            continue

        safe_team_name = safe_file_name(team_name)
        temp_path      = os.path.join(SAVE_DIR, f"temp_{safe_team_name}.xlsx")
        team_data.to_excel(temp_path, index=False, engine="openpyxl")

        wb = load_workbook(temp_path)
        ws = wb.active
        set_worksheet_style(ws, ws.max_row, ws.max_column)
        total_fee = add_summary_area(ws, team_data, team_name)
        wb.save(temp_path)
        wb.close()

        try:
            team_data["业务时间"] = pd.to_datetime(team_data["业务时间"])
            target_time = team_data.iloc[1 if len(team_data) >= 2 else 0]["业务时间"]
            year_month  = target_time.strftime("%Y-%m")
        except Exception:
            year_month = "0000-00"

        if abs(total_fee) < 1e-6 and team_name != SPECIAL_TEAM:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            print(f"【跳过】{team_name}：合计应付为0")
            continue

        fee_str       = f"{total_fee:.2f}"
        new_file_name = f"{fee_str}-{safe_team_name}_{year_month}_快递加收费.xlsx"
        new_path      = os.path.join(SAVE_DIR, new_file_name)
        if os.path.exists(new_path):
            os.remove(new_path)
        os.rename(temp_path, new_path)
        print(f"✅ 生成完成：{new_file_name} | 单据数：{len(team_data)} | 合计：{total_fee}")


# ===================== 程序入口 =====================
def main():
    print("=" * 65)
    print("        客户对账账单拆分程序 V2.4 稳定版")
    print("=" * 65)
    try:
        ensure_folder(SAVE_DIR)
        load_team_config()
        source_df = load_source_data()
        split_by_group(source_df)
        print("\n" + "=" * 65)
        print("🎉 全部任务执行完毕")
        print("=" * 65)
    except Exception as e:
        print(f"\n❌ 运行异常：{str(e)}")


if __name__ == "__main__":
    main()